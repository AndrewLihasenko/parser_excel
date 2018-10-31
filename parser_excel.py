""" Script reads data from the table Excel """

import openpyxl
import os

try:
    wb = openpyxl.load_workbook(
        r'\\migremont.local\Public\Migremont\Тех. Отдел\Конструкт. бюро\Развертки для лазера\Вопросы по разверткам.xlsx')
    ws = wb['Лист1']

    result = ''
    for row in ws.iter_rows(min_row=1, max_col=6):
        for cell in row:
            if cell.value is None:
                cells_None = [cell.coordinate]
                for val in cells_None:
                    cell_B = ws.cell(row=int(val[1:]), column=2).value
                    cell_C = ws.cell(row=int(val[1:]), column=3).value
                    cell_D = ws.cell(row=int(val[1:]), column=4).value
                    if cell_B is not None and cell_D is None:
                        list_cell_B = [cell_B]
                        for n in list_cell_B:
                            if n not in result:
                                result = ''.join(n) + ' - ' + ''.join(cell_C)
                                print(result)
                                cmd = "msg * /server:192.168.0.25 Найдены вопросы по разверткам: "
                                os.system(cmd + ' ' + result)
except FileNotFoundError:
    cmd = "msg * /server:192.168.0.25 Файл 'Вопросы по разверткам.xlsx' не найден!"
    os.system(cmd)
    print("Файл 'Вопросы по разверткам.xlsx' не найден!")

    